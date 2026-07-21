#!/usr/bin/env python3
"""Compile the canonical Bridge CSV registries into the v3 XLSX workbook."""
from __future__ import annotations
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
OUTPUT = ROOT / "T4H_Bridge_Autonomous_Repair_and_Distributed_Work_v3.xlsx"

SHEETS = [
    ("Error_Fix_Backlog", "error_fix_backlog.csv"),
    ("Job_Pipeline", "job_pipeline.csv"),
    ("Worker_Registry", "worker_registry.csv"),
    ("Worker_Contract", "worker_contract.csv"),
    ("Job_Contract", "job_contract.csv"),
    ("Handover_Scratchpad", "handover_scratchpad.csv"),
    ("Tool_LLM_Interconnect", "tool_llm_interconnect.csv"),
    ("Auto_Fix_Playbooks", "auto_fix_playbooks.csv"),
    ("Known_Blockers", "known_blockers.csv"),
    ("Data_Dictionary", "data_dictionary.csv"),
]

TITLE = "T4H Bridge Autonomous Repair & Distributed Work Pipeline"
SUBTITLE = (
    "Detect failures, create executable jobs, route humans and machines, repair within authority, "
    "verify by readback and receipts, preserve scratchpad state, and continue work across tools and LLMs."
)

def read_csv(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.reader(f))

def style_sheet(ws) -> None:
    dark = PatternFill("solid", fgColor="0F172A")
    header = PatternFill("solid", fgColor="334155")
    for cell in ws[1]:
        cell.fill = dark
        cell.font = Font(color="FFFFFF", bold=True, size=14)
    for cell in ws[4]:
        cell.fill = header
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A5"
    for column in ws.columns:
        letter = column[0].column_letter
        width = min(max(12, max(len(str(c.value or "")) for c in column) + 2), 32)
        ws.column_dimensions[letter].width = width

def add_registry_sheet(wb: Workbook, sheet_name: str, csv_name: str) -> None:
    rows = read_csv(DATA / csv_name)
    ws = wb.create_sheet(sheet_name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(rows[0]))
    ws.cell(1, 1, sheet_name.replace("_", " "))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(rows[0]))
    ws.cell(2, 1, SUBTITLE)
    for row in rows:
        ws.append(row)
    ref = f"A4:{ws.cell(ws.max_row, ws.max_column).coordinate}"
    table = Table(displayName=sheet_name.replace("_", "") + "Table", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)
    style_sheet(ws)

def add_dashboard(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Dashboard"
    ws.merge_cells("A1:J1")
    ws["A1"] = TITLE
    ws.merge_cells("A2:J2")
    ws["A2"] = SUBTITLE
    ws.append([])
    ws.append(["Metric", "Formula / Value"])
    ws.append(["Backlog items", "=COUNTA('Error_Fix_Backlog'!A5:A500)"])
    ws.append(["P0 items", '=COUNTIF(\'Error_Fix_Backlog\'!D5:D500,"P0")'])
    ws.append(["Auto-fix eligible", '=COUNTIF(\'Error_Fix_Backlog\'!K5:K500,"AUTO")'])
    ws.append(["Blocked", '=COUNTIF(\'Error_Fix_Backlog\'!E5:E500,"BLOCKED")'])
    ws.append(["Open jobs", '=COUNTA(\'Job_Pipeline\'!A5:A500)-COUNTIF(\'Job_Pipeline\'!H5:H500,"DONE")'])
    ws.append(["Workers registered", "=COUNTA('Worker_Registry'!A5:A500)"])
    style_sheet(ws)

def main() -> int:
    wb = Workbook()
    add_dashboard(wb)
    for sheet_name, csv_name in SHEETS:
        add_registry_sheet(wb, sheet_name, csv_name)
    wb.save(OUTPUT)
    print(OUTPUT)
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
