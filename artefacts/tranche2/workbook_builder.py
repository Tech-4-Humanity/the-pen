from openpyxl import Workbook

sheets = [
    '00_README','01_ORGANISATION','02_OWNERSHIP','03_BENEFICIAL_OWNERS',
    '04_STAFF','05_SERVICES','06_CUSTOMER_PROFILE','07_GEOGRAPHIC_EXPOSURE',
    '08_TRANSACTION_PROFILE','09_MONEY_MOVEMENT','10_TECHNOLOGY',
    '11_EXISTING_CONTROLS','12_RISK_WORKSHOP','13_EVIDENCE_CHECKLIST',
    '14_TRAINING_MATRIX','15_GAP_ANALYSIS','16_GENERATED_OUTPUTS',
    '17_ANNUAL_CALENDAR','18_INDEPENDENT_REVIEW','19_DECLARATIONS',
    '20_EXECUTIVE_SIGNOFF'
]

headers = {
    '01_ORGANISATION':['legal_name','abn','acn','entity_type','employees','turnover_band'],
    '02_OWNERSHIP':['owner_name','role','ownership_pct','pep','country'],
    '04_STAFF':['name','role','office','manager','training_level','police_check'],
    '14_TRAINING_MATRIX':['role','level','hours','annual_refresh_required'],
    '15_GAP_ANALYSIS':['obligation','status','evidence','action_required']
}

wb = Workbook()
wb.active.title = sheets[0]
wb.active['A1'] = 'CDAP Master Workbook v1.0'

for sheet_name in sheets[1:]:
    ws = wb.create_sheet(sheet_name)
    for col_index, header in enumerate(headers.get(sheet_name, ['field','value','notes']), 1):
        ws.cell(1, col_index, header)

wb.save('CDAP_Master_Workbook_v1.xlsx')
print('Generated CDAP_Master_Workbook_v1.xlsx')
